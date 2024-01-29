from django.shortcuts import render, HttpResponse
from .models import rbi_log
import json

from django.http import JsonResponse,request
from django.core.serializers import serialize
from django.utils import timezone
from datetime import timedelta
from django.http import HttpResponseBadRequest

from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET
from datetime import date,datetime, timedelta
from django.shortcuts import get_object_or_404
from calendar import monthrange
from django.db.models import Q
from datetime import datetime, timedelta
from .forms import DateRangeForm
from django.core.exceptions import ObjectDoesNotExist
from django.core.exceptions import ValidationError
from django import forms
from django.db import models

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
import configparser
import os
from django.shortcuts import render
import sys


from django.db.models import F

# Add the directory containing probe_agile_data to the Python path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))





def get_status_from_config(source_name):
    config_path = get_config_path(source_name)
    config = configparser.ConfigParser()
    config.read(config_path)
    return config.get(source_name, 'status')

def get_config_path(source_name):
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    config_folder = os.path.join(base_dir, 'config')
    config_file = f'config_{source_name.lower()}.ini'
    return os.path.join(config_folder, config_file)


# Function to get a unique Sr_no value
def get_unique_sr_no():
    max_sr_no = rbi_log.objects.using('rbi').aggregate(models.Max('Sr_no'))['Sr_no__max']
    return max_sr_no + 1 if max_sr_no is not None else 1

def rbinewhome(request):
    # Get the current date and the date 6 days ago
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=6)
    
    # Query the database for rbi_fema , rbi_ecb ,rbi_odi, startupindia data within the last 7 days
    fema_data = rbi_log.objects.using('rbi').filter(source_name='rbi_fema', date_of_scraping__date__range=[start_date, end_date]).order_by('-date_of_scraping')
    ecb_data = rbi_log.objects.using('rbi').filter(source_name='rbi_ecb', date_of_scraping__date__range=[start_date, end_date]).order_by('-date_of_scraping')
    odi_data = rbi_log.objects.using('rbi').filter(source_name='rbi_odi', date_of_scraping__date__range=[start_date, end_date]).order_by('-date_of_scraping')
    
    startupindia_data = rbi_log.objects.using('rbi').filter(source_name='startupindia', date_of_scraping__date__range=[start_date, end_date]).order_by('-date_of_scraping')
   
    
    # Try to get the latest total_record_count for rbi_fema and rbi_ecb, rbi_odi, startupindia
    try:
        fema_latest_entry = rbi_log.objects.using('rbi').filter(source_name='rbi_fema').latest('date_of_scraping')
        # Set fema_latest_count to 0 for success, '-' for failure, or '-' if total_record_count is None
        fema_latest_count = fema_latest_entry.total_record_count if fema_latest_entry.total_record_count is not None else "0" if fema_latest_entry.script_status == 'Success' else "-"
    except ObjectDoesNotExist:
        # Set fema_latest_count to '-' if there is no latest entry
        fema_latest_count = "-"

    try:
        ecb_latest_entry = rbi_log.objects.using('rbi').filter(source_name='rbi_ecb').latest('date_of_scraping')
        # Set ecb_latest_count to 0 for success, '-' for failure, or '-' if total_record_count is None
        ecb_latest_count = ecb_latest_entry.total_record_count if ecb_latest_entry.total_record_count is not None else "0" if ecb_latest_entry.script_status == 'Success' else "-"
    except ObjectDoesNotExist:
        # Set ecb_latest_count to '-' if there is no latest entry
        ecb_latest_count = "-"
        
    try:
        odi_latest_entry = rbi_log.objects.using('rbi').filter(source_name='rbi_odi').latest('date_of_scraping')
        odi_latest_count = odi_latest_entry.total_record_count if odi_latest_entry.total_record_count is not None else "0" if odi_latest_entry.script_status == 'Success' else "-"
    except ObjectDoesNotExist:
        odi_latest_count = "-"
   
    try:
        startupindia_latest_entry = rbi_log.objects.using('rbi').filter(source_name='startupindia').latest('date_of_scraping')
        startupindia_latest_count = startupindia_latest_entry.total_record_count if startupindia_latest_entry.total_record_count is not None else "0" if startupindia_latest_entry.script_status == 'Success' else "-"
    except ObjectDoesNotExist:
        startupindia_latest_count = "-"

        
    data_list = []
    
    # Iterate over the last 7 days
    for date in (end_date - timedelta(days=i) for i in range(7)):
        # Get the rbi_fema entry and rbi_ecb entry ,rbi_odi,startupindia for the current date
        fema_entry = fema_data.filter(date_of_scraping__date=date).first()
        ecb_entry = ecb_data.filter(date_of_scraping__date=date).first()
        odi_entry = odi_data.filter(date_of_scraping__date=date).first()
        
        startupindia_entry = startupindia_data.filter(date_of_scraping__date=date).first()
        
        # Set fema_data_available and fema_data_scraped to '0' for success, 'NA' for failure, or '-' if data is None
        fema_data_available = fema_entry.data_available if fema_entry is not None and fema_entry.data_available is not None else "0" if fema_entry is not None and fema_entry.script_status == 'Success' else "NA" if fema_entry is not None and fema_entry.script_status == 'Failure' else "-"
        fema_data_scraped = fema_entry.data_scraped if fema_entry is not None and fema_entry.data_scraped is not None else "0" if fema_entry is not None and fema_entry.script_status == 'Success' else "NA" if fema_entry is not None and fema_entry.script_status == 'Failure' else "-"
       
        # Set ecb_data_available and ecb_data_scraped to '0' for success, 'NA' for failure, or '-' if data is None
        ecb_data_available = ecb_entry.data_available if ecb_entry is not None and ecb_entry.data_available is not None else "0" if ecb_entry is not None and ecb_entry.script_status == 'Success' else "NA" if ecb_entry is not None and ecb_entry.script_status == 'Failure' else "-"
        ecb_data_scraped = ecb_entry.data_scraped if ecb_entry is not None and ecb_entry.data_scraped is not None else "0" if ecb_entry is not None and ecb_entry.script_status == 'Success' else "NA" if ecb_entry is not None and ecb_entry.script_status == 'Failure' else "-"
       
        
        odi_data_available = odi_entry.data_available if odi_entry is not None and odi_entry.data_available is not None else "0" if odi_entry is not None and odi_entry.script_status == 'Success' else "NA" if odi_entry is not None and odi_entry.script_status == 'Failure' else "-"
        odi_data_scraped = odi_entry.data_scraped if odi_entry is not None and odi_entry.data_scraped is not None else "0" if odi_entry is not None and odi_entry.script_status == 'Success' else "NA" if odi_entry is not None and odi_entry.script_status == 'Failure' else "-"
        
        startupindia_data_available = startupindia_entry.data_available if startupindia_entry is not None and startupindia_entry.data_available is not None else "0" if startupindia_entry is not None and startupindia_entry.script_status == 'Success' else "NA" if startupindia_entry is not None and startupindia_entry.script_status == 'Failure' else "-"
        startupindia_data_scraped = startupindia_entry.data_scraped if startupindia_entry is not None and startupindia_entry.data_scraped is not None else "0" if startupindia_entry is not None and startupindia_entry.script_status == 'Success' else "NA" if startupindia_entry is not None and startupindia_entry.script_status == 'Failure' else "-"
       
         
        fema_status = fema_entry.script_status if fema_entry is not None else 'N/A'
        ecb_status = ecb_entry.script_status if ecb_entry is not None else 'N/A'
        odi_status = odi_entry.script_status if odi_entry is not None else 'N/A'
        startupindia_status = startupindia_entry.script_status if startupindia_entry is not None else 'N/A'
        
        
        fema_reason = fema_entry.failure_reason if fema_entry is not None else None
        ecb_reason = ecb_entry.failure_reason if ecb_entry is not None else None
        odi_reason = odi_entry.failure_reason if odi_entry is not None else None
        startupindia_reason = startupindia_entry.failure_reason if startupindia_entry is not None else None
        
        

        # Determine the color based on status and reason
        fema_color = (
            'green' if fema_status == 'Success' else
            'orange' if fema_status == 'Failure' and '204' in str(fema_reason) else
            'red' if fema_status == 'Failure' else
            'black'
        )
  
        ecb_color = (
            'green' if ecb_status == 'Success' else
            'orange' if ecb_status == 'Failure' and '204' in str(ecb_reason) else
            'red' if ecb_status == 'Failure' else
            'black'
        )  
        
        odi_color = (
            'green' if odi_status == 'Success' else
            'orange' if odi_status == 'Failure' and '204' in str(odi_reason) else
            'red' if odi_status == 'Failure' else
            'black'
        )
        
        startupindia_color = (
            'green' if startupindia_status == 'Success' else
            'orange' if startupindia_status == 'Failure' and '204' in str(startupindia_reason) else
            'red' if startupindia_status == 'Failure' else
            'black'
        )
        
        
        # Append data to the data_list for rendering in HTML
        data_list.append({
            'Date': date.strftime('%d-%m-%Y'),
            'FEMA_Data_Available': fema_data_available,
            'FEMA_Data_Scraped': fema_data_scraped,
            'FEMA_Color': fema_color,
            'ECB_Data_Available': ecb_data_available,
            'ECB_Data_Scraped': ecb_data_scraped,
            'ECB_Color': ecb_color,
            'ODI_Data_Available': odi_data_available,
            'ODI_Data_Scraped': odi_data_scraped,
            'ODI_Color': odi_color,
            'STARTUPINDIA_Data_Available' : startupindia_data_available,
            'STARTUPINDIA_Data_Scraped' :startupindia_data_scraped,
            'STARTUPINDIA_Color':startupindia_color,
        })
    
    
    
    fema_status = get_status_from_config('rbi_fema')
    ecb_status = get_status_from_config('rbi_ecb')
    odi_status = get_status_from_config('rbi_odi')
    startupindia_status = get_status_from_config('startupindia')

    status_color_mapping = {
        'Active': 'green',
        'Hibernated': 'orange',
        'Inactive': 'red',
    }

    fema_status_color = status_color_mapping.get(fema_status, '')
    ecb_status_color = status_color_mapping.get(ecb_status, '')
    odi_status_color = status_color_mapping.get(odi_status, '')
    startupindia_status_color = status_color_mapping.get(startupindia_status, '')
    
    unique_sr_no = get_unique_sr_no() 
    
    
    # rbi_log.objects.using('rbi').filter(source_name='rbi_fema').update_or_create(defaults={'source_status': fema_status, 'Sr_no': unique_sr_no})
    # rbi_log.objects.using('rbi').filter(source_name='rbi_ecb').update_or_create(defaults={'source_status': ecb_status, 'Sr_no': unique_sr_no})
    # rbi_log.objects.using('rbi').filter(source_name='rbi_odi').update_or_create(defaults={'source_status': odi_status, 'Sr_no': unique_sr_no})
    # rbi_log.objects.using('rbi').filter(source_name='startupindia').update_or_create(defaults={'source_status': startupindia_status, 'Sr_no': unique_sr_no})
    
         
    # Update or create for rbi_fema
    rbi_fema_entry, created = rbi_log.objects.using('rbi').filter(source_name='rbi_fema', Sr_no=unique_sr_no ).update_or_create(source_name='rbi_fema', Sr_no=unique_sr_no,  defaults={'source_status': fema_status} )

    # If entry is not created, update the Sr_no
    if not created:
        rbi_fema_entry.Sr_no = unique_sr_no
        rbi_fema_entry.save()
    
    
    # Get a new unique Sr_no for the next update or create operation
    unique_sr_no = get_unique_sr_no()
        
    # Update or create for rbi_ecb
    rbi_ecb_entry,created = rbi_log.objects.using('rbi').filter(source_name='rbi_ecb', Sr_no=unique_sr_no ).update_or_create(source_name='rbi_ecb',  Sr_no=unique_sr_no, defaults={'source_status': ecb_status})

    # If entry is not created, update the Sr_no
    if not created:
        rbi_ecb_entry.Sr_no = unique_sr_no()
        rbi_ecb_entry.save()
        
        
    # Get a new unique Sr_no for the next update or create operation
    unique_sr_no = get_unique_sr_no()

    # Update or create for rbi_odi
    rbi_odi_entry, created= rbi_log.objects.using('rbi').filter(source_name='rbi_odi', Sr_no=unique_sr_no ).update_or_create(source_name='rbi_odi', Sr_no=unique_sr_no,  defaults={'source_status': odi_status})


    # If entry is not created, update the Sr_no
    if not created:
        rbi_odi_entry.Sr_no = unique_sr_no()
        rbi_odi_entry.save()
    
    
    # Get a new unique Sr_no for the next update or create operation
    unique_sr_no = get_unique_sr_no()
    
    # Update or create for startupindia
    startupindia_entry, created = rbi_log.objects.using('rbi').filter(source_name='startupindia', Sr_no=unique_sr_no ).update_or_create(source_name='startupindia', Sr_no=unique_sr_no,  defaults={'source_status': startupindia_status})

    # If entry is not created, update the Sr_no
    if not created:
        startupindia_entry.Sr_no = unique_sr_no()
        startupindia_entry.save()
   
    
    # Get a new unique Sr_no for the next update or create operation
    unique_sr_no = get_unique_sr_no()
    
    
    
    # rbi_log.objects.using('rbi').filter(source_name='rbi_fema').update_or_create(source_name='rbi_fema', defaults={'source_status': fema_status, 'Sr_no': unique_sr_no})
    # rbi_log.objects.using('rbi').filter(source_name='rbi_ecb').update_or_create(source_name='rbi_ecb', defaults={'source_status': ecb_status, 'Sr_no': unique_sr_no})
    # rbi_log.objects.using('rbi').filter(source_name='rbi_odi').update_or_create(source_name='rbi_odi', defaults={'source_status': odi_status, 'Sr_no': unique_sr_no})
    # rbi_log.objects.using('rbi').filter(source_name='startupindia').update_or_create(source_name='startupindia', defaults={'source_status': startupindia_status, 'Sr_no': unique_sr_no})
        
    
    # Prepare the context to be passed to the HTML template
    context = {
        'data_list': data_list, 
        'fema_latest_count': fema_latest_count, 
        'ecb_latest_count': ecb_latest_count, 
        'odi_latest_count': odi_latest_count, 
        'startupindia_latest_count': startupindia_latest_count,
        'fema_status': fema_status,
        'ecb_status': ecb_status,
        'odi_status': odi_status,
        'startupindia_status': startupindia_status,
        'fema_status_color': fema_status_color,
        'ecb_status_color': ecb_status_color,
        'odi_status_color': odi_status_color,
        'startupindia_status_color': startupindia_status_color,
        
    }
    
    # Render the HTML template with the context
    return render(request, 'fema/grid.html', context)



  
def rbiget_data_for_popup1(request, source_name):
    today_date = timezone.now().date()
    data = rbi_log.objects.using('rbi').filter(source_name=source_name, date_of_scraping__date=today_date).first()

    if data:
        
         # Replace None values with hyphen
        data_scraped = data.data_scraped if data.data_scraped is not None else "0"
        failure_reason = data.failure_reason if data.failure_reason is not None else "-"

        response_data = {
            'source_name': data.source_name,
            'script_status': data.script_status,
         
            'data_scraped': data_scraped,
          
          
            'failure_reason': failure_reason,
            
            'date_of_scraping': data.date_of_scraping.strftime('%d-%m-%Y'),
        }
        print(f"Today's Date: {today_date}")
        print(f"Source Name: {source_name}")

        return HttpResponse(json.dumps(response_data), content_type="application/json")
    else:
        return HttpResponse(status=404)
    



###################################################################################################################################################




def get_config_path(source_name):
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    config_folder = os.path.join(base_dir, 'config')
    config_file = f'config_{source_name.lower()}.ini'
    return os.path.join(config_folder, config_file)


def read_config(source_name):
    config_path = get_config_path(source_name)
    
    config = configparser.ConfigParser()
    config.read(config_path)

    return config.get(source_name, 'status')




"""Format a date to a string in 'dd-mm-YYYY' format."""
def format_date(date):
    return date.strftime('%d-%m-%Y') if date else ''

"""Get default start and end dates for a date range of past 7 days."""
def get_default_start_end_dates():
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=6)  # Default to past 7 days
    return start_date, end_date

"""Get start and end dates for a date range of past 15 days."""
def get_past_15_days():
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=14)
    return start_date, end_date

"""Get start and end dates for a date range of past month (30 days)."""
def get_past_month():
    today = datetime.now().date()
    end_date = today
    start_date = today - timedelta(days=29)  # Modify to get data for the past 30 days, including today
    return start_date, end_date

#  This code is for past month results .for example current month is december past month results is november month.
# def get_last_month():
#     today = datetime.now().date()
#     end_date = today.replace(day=1) - timedelta(days=1)
#     start_date = end_date.replace(day=1)
#     return start_date, end_date

"""Get color based on script_status and failure_reason."""
def get_status_color(script_status, failure_reason):
    if script_status == 'Success':
        return 'green'
    elif script_status == 'Failure' and '204' in str(failure_reason):
        return 'orange'
    else:
        return 'red'



def filter_data(request, source_name):
    """Filter and process data based on date range and source name."""
    form = DateRangeForm(request.GET)
    
    # Default values for start_date and end_date
    start_date, end_date = get_default_start_end_dates()
    
    if form.is_valid():
        date_range = form.cleaned_data.get('date_range')
        if date_range == 'past_7_days':
            start_date, end_date = get_default_start_end_dates()
        elif date_range == 'past_15_days':
            start_date, end_date = get_past_15_days()
        elif date_range == 'past_month':
            start_date, end_date = get_past_month()
        elif date_range == 'custom':
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')
            if start_date and end_date:
                date_difference = end_date - start_date
                if date_difference.days > 60:
                    # Adjust end_date if it's more than 60 days from start_date
                    end_date = start_date + timedelta(days=60)
        else:
            start_date, end_date = get_default_start_end_dates()

    # Adjust end_date to cover the entire day
    # end_date = end_date + timedelta(days=1)
    
    
    # Query the database with the adjusted date range
    data = rbi_log.objects.using('rbi').filter(
         date_of_scraping__date__range=[start_date, end_date],
        source_name=source_name
    ).order_by('-date_of_scraping') # Sort data in descending order based on date
    
    # Process and format the data
    formatted_data = []
    for item in data:
        formatted_date = format_date(item.date_of_scraping)
        status_color = get_status_color(item.script_status, item.failure_reason)
        
        # Replace None values with hyphen .none value is a null value from database.
        data_available = item.data_available if item.data_available is not None else "0"
        data_scraped = item.data_scraped if item.data_scraped is not None else "0"
        failure_reason = item.failure_reason if item.failure_reason is not None else "-"
        
        formatted_data.append({
            'source_name': item.source_name,
            'script_status': item.script_status,
            'failure_reason': failure_reason,
            'data_available': data_available,
            'data_scraped': data_scraped,
            'date_of_scraping': formatted_date,
            'status_color': status_color,
        })
        
    
    # Handle export to Excel functionality
    if 'download' in request.GET:
        date_range = request.GET.get('date_range', 'past_7_days')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        # Call the export_to_excel function with the selected date range
        return export_to_excel(request, formatted_data, date_range, start_date, end_date, source_name)
    
    
    # Read status for the current source_name
    current_status = read_config(source_name)
    
     # Set color based on the current_status
    if current_status == 'Active':
        status_color = 'greenyellow'
    elif current_status == 'Hibernated':
        status_color = 'orange'
    elif current_status == 'Inactive':
        status_color = 'red'
    else:
        status_color = 'black'  # Set a default color for unknown statuses

    
    # Prepare context for rendering the template
    context = {
        'form': form,
        'data': formatted_data,
        'start_date': format_date(start_date),
        'end_date': format_date(end_date),
        'past_15_days': (format_date(get_past_15_days()[0]), format_date(get_past_15_days()[1])),
        'last_month': (format_date(get_past_month()[0]), format_date(get_past_month()[1])),
        'source_name': source_name,
        'current_status': current_status,
        'status_color': status_color,
    }

    return render(request, 'fema/gridfilter.html', context)




def rbinewfema_datefilter(request):
    return filter_data(request, 'rbi_fema')

def rbinewecb_datefilter(request):
    return filter_data(request, 'rbi_ecb')

def rbinewodi_datefilter(request):
    return filter_data(request, 'rbi_odi')

def rbinewstartupindia_datefilter(request):
    return filter_data(request, 'startupindia')






def export_to_excel(request, data, date_range, start_date, end_date, source_name):
    """Export data to an Excel file."""
    if date_range == 'past_15_days':
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=14)
    elif date_range == 'past_month':
        today = datetime.now().date()
        end_date = today
        start_date = today - timedelta(days=29)
    elif date_range == 'custom':
        # Add logic to handle custom view start_date and end_date
        if start_date and end_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            # Default to past 7 days if no specific range is selected
            end_date = datetime.now().date()
            start_date = end_date - timedelta(days=6)
    else:
        # Default to past 7 days if no specific range is selected
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=6)

    excel_data = rbi_log.objects.using('rbi').filter(date_of_scraping__date__range=[start_date, end_date], source_name=source_name).order_by('-date_of_scraping')
    
    # Generate a dynamic filename based on source name and date range
    filename = f"{source_name}_{start_date}_{end_date}.xlsx"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Generate Excel file
    workbook = Workbook()
    worksheet = workbook.active
    
    # Add headers to the worksheet
    header_font = Font(bold=True)
    headers = ['Source Name', 'Status',  '#Records Available', '#Records Scraped', 'Failure Reason', 'Scraped On']

    for col_num, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
    
    # Populate data in the worksheet
    for row_num, data_entry in enumerate(excel_data, start=2):
        worksheet.cell(row=row_num, column=1, value=data_entry.source_name)
        worksheet.cell(row=row_num, column=2, value=data_entry.script_status)
        worksheet.cell(row=row_num, column=3, value=data_entry.data_available if data_entry.data_available is not None else "0")  
        worksheet.cell(row=row_num, column=4, value=data_entry.data_scraped if data_entry.data_scraped is not None else "0")
        worksheet.cell(row=row_num, column=5, value=data_entry.failure_reason or "-")
        worksheet.cell(row=row_num, column=6, value=format_date(data_entry.date_of_scraping))
        # worksheet.cell(row=row_num, column=7, value=get_status_color(data_entry.script_status, data_entry.failure_reason))
    
    # Save the workbook and prepare the response for download
    workbook.save(response)
    return response


###################################################################################################################################################

  
  
# def rbinewfema_datefilter(request):
#     form = DateRangeForm(request.GET)
    
#     # Default values for start_date and end_date
#     start_date, end_date = get_default_start_end_dates()
    
#     if form.is_valid():
#         date_range = form.cleaned_data.get('date_range')
#         if date_range == 'past_7_days':
#             start_date, end_date = get_default_start_end_dates()
#         elif date_range == 'past_15_days':
#             start_date, end_date = get_past_15_days()
#         elif date_range == 'last_month':
#             start_date, end_date = get_last_month()
#         elif date_range == 'custom':
#             start_date = form.cleaned_data.get('start_date')
#             end_date = form.cleaned_data.get('end_date')
#             if start_date and end_date:
#                 date_difference = end_date - start_date
#                 if date_difference.days > 60:
#                     # Adjust end_date if it's more than 60 days from start_date
#                     end_date = start_date + timedelta(days=60)
#         else:
#             start_date, end_date = get_default_start_end_dates()

#     data = rbi_log.objects.using('rbi').filter(
#         date_of_scraping__date__range=[start_date, end_date],
#         source_name='rbi_fema'
#     )

#     formatted_data = []
#     for item in data:
#         formatted_date = format_date(item.date_of_scraping)
#         status_color = get_status_color(item.script_status, item.failure_reason)
        
#         # Replace None values with hyphen
#         data_available = item.data_available if item.data_available is not None else "0"
#         data_scraped = item.data_scraped if item.data_scraped is not None else "0"
#         failure_reason = item.failure_reason if item.failure_reason is not None else "-"
        
#         formatted_data.append({
#             'source_name': item.source_name,
#             'script_status': item.script_status,
#             'failure_reason': failure_reason,
#             'data_available': data_available,
#             'data_scraped': data_scraped,
#             'date_of_scraping': formatted_date,
#             'status_color': status_color,
#         })

#     context = {
#         'form': form,
#         'data': formatted_data,
#         'start_date': format_date(start_date),
#         'end_date': format_date(end_date),
#         'past_15_days': (format_date(get_past_15_days()[0]), format_date(get_past_15_days()[1])),
#         'last_month': (format_date(get_last_month()[0]), format_date(get_last_month()[1])),
#         'table_name_filter': 'rbi_fema',
#         'source_name':'rbi_fema',
#     }

#     return render(request, 'fema/download_data.html', context)



# def rbinewecb_datefilter(request):
#     form = DateRangeForm(request.GET)
    
    
#     # Default values for start_date and end_date
#     start_date, end_date = get_default_start_end_dates()
    
#     if form.is_valid():
#         date_range = form.cleaned_data.get('date_range')
#         if date_range == 'past_7_days':
#             start_date, end_date = get_default_start_end_dates()
#         elif date_range == 'past_15_days':
#             start_date, end_date = get_past_15_days()
#         elif date_range == 'last_month':
#             start_date, end_date = get_last_month()
#         elif date_range == 'custom':
#             start_date = form.cleaned_data.get('start_date')
#             end_date = form.cleaned_data.get('end_date')
#             if start_date and end_date:
#                 date_difference = end_date - start_date
#                 if date_difference.days > 60:
#                     # Adjust end_date if it's more than 60 days from start_date
#                     end_date = start_date + timedelta(days=60)
#         else:
#             start_date, end_date = get_default_start_end_dates()

#     # Adjust end_date to cover the entire day
#     # end_date = end_date + timedelta(days=1)

#     data = rbi_log.objects.using('rbi').filter(
#          date_of_scraping__date__range=[start_date, end_date],
#         source_name='rbi_ecb'
#     )

#     formatted_data = []
#     for item in data:
#         formatted_date = format_date(item.date_of_scraping)
#         status_color = get_status_color(item.script_status, item.failure_reason)
        
#         # Replace None values with hyphen
#         data_available= item.data_available if item.data_available is not None else "0"
#         data_scraped = item.data_scraped if item.data_scraped is not None else "0"
#         failure_reason = item.failure_reason if item.failure_reason is not None else "-"
        
#         formatted_data.append({
#             'source_name': item.source_name,
#             'script_status': item.script_status,
#             'failure_reason':failure_reason,
#             'data_available':data_available,
#             'data_scraped': data_scraped,
#             'date_of_scraping': formatted_date,
#             'status_color': status_color,
#         })

#     context = {
#         'form': form,
#         'data': formatted_data,
#         'start_date': format_date(start_date),
#         'end_date': format_date(end_date),
#         'past_15_days': (format_date(get_past_15_days()[0]), format_date(get_past_15_days()[1])),
#         'last_month': (format_date(get_last_month()[0]), format_date(get_last_month()[1])),
#         'table_name_filter': 'rbi_ecb',
#         'source_name':'rbi_ecb',
#     }

#     return render(request, 'fema/download_data.html', context)



# def rbinewodi_datefilter(request):
#     form = DateRangeForm(request.GET)
    
    
#     # Default values for start_date and end_date
#     start_date, end_date = get_default_start_end_dates()
    
#     if form.is_valid():
#         date_range = form.cleaned_data.get('date_range')
#         if date_range == 'past_7_days':
#             start_date, end_date = get_default_start_end_dates()
#         elif date_range == 'past_15_days':
#             start_date, end_date = get_past_15_days()
#         elif date_range == 'last_month':
#             start_date, end_date = get_last_month()
#         elif date_range == 'custom':
#             start_date = form.cleaned_data.get('start_date')
#             end_date = form.cleaned_data.get('end_date')
#             if start_date and end_date:
#                 date_difference = end_date - start_date
#                 if date_difference.days > 60:
#                     # Adjust end_date if it's more than 60 days from start_date
#                     end_date = start_date + timedelta(days=60)
#         else:
#             start_date, end_date = get_default_start_end_dates()

#     # Adjust end_date to cover the entire day
#     # end_date = end_date + timedelta(days=1)

#     data = rbi_log.objects.using('rbi').filter(
#          date_of_scraping__date__range=[start_date, end_date],
#         source_name='rbi_odi'
#     )

#     formatted_data = []
#     for item in data:
#         formatted_date = format_date(item.date_of_scraping)
#         status_color = get_status_color(item.script_status, item.failure_reason)
        
#         # Replace None values with hyphen
#         data_available= item.data_available if item.data_available is not None else "0"
#         data_scraped = item.data_scraped if item.data_scraped is not None else "0"
#         failure_reason = item.failure_reason if item.failure_reason is not None else "-"
        
#         formatted_data.append({
#             'source_name': item.source_name,
#             'script_status': item.script_status,
#             'failure_reason':failure_reason,
#             'data_available':data_available,
#             'data_scraped': data_scraped,
#             'date_of_scraping': formatted_date,
#             'status_color': status_color,
#         })

#     context = {
#         'form': form,
#         'data': formatted_data,
#         'start_date': format_date(start_date),
#         'end_date': format_date(end_date),
#         'past_15_days': (format_date(get_past_15_days()[0]), format_date(get_past_15_days()[1])),
#         'last_month': (format_date(get_last_month()[0]), format_date(get_last_month()[1])),
#         'table_name_filter': 'rbi_odi',
#     }

#     return render(request, 'fema/gridfilter.html', context)


   
# def rbinewstartupindia_datefilter(request):
#     form = DateRangeForm(request.GET)
    
#      # Default values for start_date and end_date
#     start_date, end_date = get_default_start_end_dates()
    
#     if form.is_valid():
#         date_range = form.cleaned_data.get('date_range')
#         if date_range == 'past_7_days':
#             start_date, end_date = get_default_start_end_dates()
#         elif date_range == 'past_15_days':
#             start_date, end_date = get_past_15_days()
#         elif date_range == 'last_month':
#             start_date, end_date = get_last_month()
#         elif date_range == 'custom':
#             start_date = form.cleaned_data.get('start_date')
#             end_date = form.cleaned_data.get('end_date')
#             if start_date and end_date:
#                 date_difference = end_date - start_date
#                 if date_difference.days > 60:
#                     # Adjust end_date if it's more than 60 days from start_date
#                     end_date = start_date + timedelta(days=60)
#         else:
#             start_date, end_date = get_default_start_end_dates()

#     # Adjust end_date to cover the entire day
#     # end_date = end_date + timedelta(days=1)

#     data = rbi_log.objects.using('rbi').filter(
#          date_of_scraping__date__range=[start_date, end_date],
#         source_name='startupindia'
#     )

#     formatted_data = []
#     for item in data:
#         formatted_date = format_date(item.date_of_scraping)
#         status_color = get_status_color(item.script_status, item.failure_reason)
        
#          # Replace None values with hyphen
#         data_available= item.data_available if item.data_available is not None else "0"
#         data_scraped = item.data_scraped if item.data_scraped is not None else "0"
#         failure_reason = item.failure_reason if item.failure_reason is not None else "-"
        
#         formatted_data.append({
#             'source_name': item.source_name,
#             'script_status': item.script_status,
#             'failure_reason': failure_reason,
#             'data_available': data_available,
#             'data_scraped': data_scraped,
#             'date_of_scraping': formatted_date,
#             'status_color': status_color,
#         })

#     context = {
#         'form': form,
#         'data': formatted_data,
#         'start_date': format_date(start_date),
#         'end_date': format_date(end_date),
#         'past_15_days': (format_date(get_past_15_days()[0]), format_date(get_past_15_days()[1])),
#         'last_month': (format_date(get_last_month()[0]), format_date(get_last_month()[1])),
#         'table_name_filter': 'startupindia',
#     }

#     return render(request, 'fema/gridfilter.html', context)



def rbi_tab(request):
    rbi_data= rbi_log.objects.using('rbi').all()
    return render(request,'fema/index.html', {'rbi_data':rbi_data}) 





    
