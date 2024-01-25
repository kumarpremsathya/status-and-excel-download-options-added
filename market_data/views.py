from .models import scraping_log
import pandas as pd
from django.http import HttpResponse, JsonResponse,request
from django.shortcuts import render
from datetime import datetime, date ,timedelta
from django.utils import timezone 
import json
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET
from django.db.models import Case, When, Value, CharField
from openpyxl import Workbook
from openpyxl.styles import Font

def dashboard(request):
    table_names = scraping_log.objects.values_list('table_name', flat=True).distinct()
    return render(request, 'dashboard.html', {'table_names': table_names})

def table_details(request, table_name):
    amber_table_names = [
        'ace_52whl',
        'ace_Company_master',
        'ace_company_master',
        'ace_Finance_fr',
        'ace_Finance_cf',
        'ace_company_equity',
        'ace_Finance_cons_cf',
        'ace_company_equity_cons',
        'ace_finance_Quarterly',
        'ace_finance_Quarterly_Cons',
        'ace_shp',
    ]

    table_list = [
                  'ace_Company_master',
                  'bse_new_equity',
                  'ace_company_equity',
                  'ace_company_equity_cons',
                  'ace_52whl',
                  'bse_market_capital',
                  'bse_financial_PROV',
                  'bse_financial_QC',
                  'ace_finance_Quarterly',
                  'ace_finance_Quarterly_Cons',
                  'ace_Finance_cf',
                  'ace_Finance_cons_cf',
                  'ace_Finance_fr',
                  'Capitaline_standalone',
                  'Capitaline_consolidated',
                  'bse_pledge',
                  'nse_pledge_new',
                  'bse_shp',
                  'ace_shp',
                  ]
    
    # Fetch all unique table names
    table_names = scraping_log.objects.values_list('table_name', flat=True).distinct()

    table_index = {name: index for index, name in enumerate(table_list)}

    # Use Case, When, and Value to order the queryset based on the custom order
    table_names = table_names.annotate(
        custom_order=Case(
            *[When(table_name=name, then=Value(index)) for name, index in table_index.items()],
            default=Value(len(table_index)),
            output_field=CharField()
        )
    ).order_by('custom_order')
    # Check if the requested table_name is valid
    if table_name not in table_names:
        return render(request, 'table_details.html', {'error_message': f'Table {table_name} not found.'})

    # Fetch all data for the selected table
    table_data = scraping_log.objects.filter(table_name=table_name)
    

    # Order the data by trade_date in descending order
    table_data = table_data.order_by('-trade_date')

    # Handle date range selection
    time_range = request.GET.get('time_range')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    today = datetime.today().date()
    yesterday = today - timedelta(days=1)

    if time_range == '7':
        start_date = today - timedelta(days=7)
        end_date = yesterday  
    elif time_range == '15':
        start_date = today - timedelta(days=15)
        end_date = yesterday
    elif time_range == '30':
        start_date = today - timedelta(days=30)
        end_date = yesterday 
    elif from_date and to_date:
        start_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(to_date, '%Y-%m-%d').date()
    else:
        # Default to last 7 days if no specific range is selected
        start_date = today - timedelta(days=7)
        end_date = yesterday 
    # Filter data based on the date range
    if start_date is not None:
        table_data = table_data.filter(trade_date__range=[start_date,end_date])

    # Get unique failure reasons
    failure_reasons = table_data.filter(status='failure').values_list('reason', flat=True).distinct()

    # Handle date filter
    selected_date = request.GET.get('scraping_date') 
    if selected_date:
        # Convert the selected_date to the desired format
        selected_date = datetime.strptime(selected_date, '%b. %d, %Y, %I:%M %p').strftime('%Y-%m-%d')
        table_data = table_data.filter(trade_date=selected_date)

    # Create a list of all dates in the desired range
    date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)][::-1]

    # Create a list of dictionaries for each date, including '-' for dates without data
    structured_data = []

    for date_str in date_range:
        data_entry = {
            'trade_date': date_str,
            'data': table_data.filter(trade_date=date_str).first(),
            'status': '-',
        }

        if data_entry['data']:
            data_entry['status'] = data_entry['data'].status.strip()

        structured_data.append(data_entry)

    data = {
        'table_name': table_name,
        'structured_data': structured_data,
        'failure_reasons': failure_reasons,
        'table_names': table_names,
        'time_range': time_range,
        'from_date': from_date,
        'to_date': to_date,
        'scraping_date': selected_date,
        'amber_table_names': amber_table_names,
        'today': today,
        'end_date':end_date,
        'start_date':start_date,
        'yesterday': yesterday,
        'table_list':table_list,
    }
    if 'download_excel' in request.GET:
        time_range = request.GET.get('time_range')
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        
# Use the custom date range if specified, otherwise use the default logic
        if time_range == '7':
            start_date = today - timedelta(days=7)
            end_date = yesterday
        elif time_range == '15':
            start_date = today - timedelta(days=15)
            end_date = yesterday
        elif time_range == '30':
            start_date = today - timedelta(days=30)
            end_date = yesterday
        elif time_range == 'custom' and from_date and to_date:
            start_date = datetime.strptime(from_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(to_date, '%Y-%m-%d').date()
            # Default to last 7 days if no specific range is selected
        else:
            start_date = datetime.now().date() - timedelta(days=7)
            end_date = yesterday

        # Fetch all dates within the specified range
        all_dates = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

        # Fetch data for the specified table and all dates within the range
        excel_data = scraping_log.objects.filter(table_name=table_name, trade_date__in=all_dates).order_by('-trade_date')
       # Create a dictionary to store data for each date
        date_data_dict = {entry.trade_date: entry for entry in excel_data}

# Add placeholder entries for dates without data
        for date in all_dates:
            if date not in date_data_dict:
            # Create a placeholder entry with default values
                placeholder_entry = scraping_log(trade_date=date, status='No Data', no_of_data_available=0, no_of_data_scraped=0, reason='No Data')
                date_data_dict[date] = placeholder_entry

                filename = f'{table_name}_data_{start_date}_{end_date}.xlsx'
        # Create a response object with the appropriate content type for Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # Construct the filename with the table_name and time_range_str
        response['Content-Disposition'] = f'attachment; filename={filename}'
        # Create a workbook and add a worksheet
        workbook = Workbook()
        worksheet = workbook.active

        # Set font style for headers
        header_font = Font(bold=True)

        # Add headers to the worksheet
        headers = ['Table Name', 'Status','#Records Available', '#Records Scraped','Failure Reason', 'Trade date', 'Scraped on']
        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font

        # Populate data into the worksheet
        for row_num, data_entry in enumerate(excel_data, start=2):
            worksheet.cell(row=row_num, column=1, value=data_entry.table_name)
            worksheet.cell(row=row_num, column=2, value=data_entry.status)
            worksheet.cell(row=row_num, column=3, value=data_entry.no_of_data_available)
            worksheet.cell(row=row_num, column=4, value=data_entry.no_of_data_scraped)
            worksheet.cell(row=row_num, column=5, value=data_entry.reason)
            worksheet.cell(row=row_num, column=6, value=data_entry.trade_date)
            worksheet.cell(row=row_num, column=7, value=data_entry.Scraped_on)

        # Save the workbook to the response
        workbook.save(response)
        return response

    return render(request, 'table_details.html', data)

def table_details2(request):
    time_range = request.GET.get('time_range')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    today = datetime.today().date()

    if time_range == '7':
        start_date = today - timedelta(days=7)
    else:
        start_date = today - timedelta(days=7)

    yesterday = today - timedelta(days=1)

    table_list = [
        'ace_Company_master',
        'bse_new_equity',
        'ace_company_equity',
        'ace_company_equity_cons',
        'ace_52whl',
        'bse_market_capital',
        'bse_financial_PROV',
        'bse_financial_QC',
        'ace_finance_Quarterly',
        'ace_finance_Quarterly_Cons',
        'ace_Finance_cf',
        'ace_Finance_cons_cf',
        'ace_Finance_fr',
        'Capitaline_standalone',
        'Capitaline_consolidated',
        'bse_pledge',
        'nse_pledge_new',
        'bse_shp',
        'ace_shp',
    ]

    # Get unique table names for the sidebar and order them based on the table_list
    table_names = scraping_log.objects.values_list('table_name', flat=True).distinct()

    # Create a dictionary to map table names to their index in the custom order
    table_index = {name: index for index, name in enumerate(table_list)}

    # Use Case, When, and Value to order the queryset based on the custom order
    table_names = table_names.annotate(
        custom_order=Case(
            *[When(table_name=name, then=Value(index)) for name, index in table_index.items()],
            default=Value(len(table_index)),
            output_field=CharField()
        )
    ).order_by('custom_order')

    amber_table_names = [
        'ace_52whl',
        'ace_Company_master',
        'ace_company_master',
        'ace_Finance_fr',
        'ace_Finance_cf',
        'ace_company_equity',
        'ace_Finance_cons_cf',
        'ace_company_equity_cons',
        'ace_finance_Quarterly',
        'ace_finance_Quarterly_Cons',
        'ace_shp',
    ]

    failure_reasons = scraping_log.objects.filter(status='failure').values_list('reason', flat=True).distinct()

    # Initialize structured_data dictionary
    structured_data = {}

    if start_date is not None:
        date_range = [start_date + timedelta(days=i) for i in range((today - start_date).days)][::-1]
    else:
        date_range = []

    # Iterate through each table name
    for table_name in table_names:
        # Filter records for the specific table and date range
        if start_date is not None:
            records = scraping_log.objects.filter(table_name=table_name).order_by('-trade_date')
        else:
            records = scraping_log.objects.filter(table_name=table_name).order_by('-trade_date')

        # Create a dictionary to store data for each date
        date_data = {}

        # Populate the dictionary with actual data
        for date in date_range:
            record = records.filter(trade_date__exact=str(date)).first()
            if record:
                record.status_stripped = record.status.strip()
                date_data[date] = {
                    'no_of_data_available': record.no_of_data_available,
                    'no_of_data_scraped': record.no_of_data_scraped,
                    'status': record.status_stripped,
                    'Scraped_on': record.Scraped_on,
                    'Table Name': record.table_name,
                    'Reason': record.reason,
                    'Trade Date': record.trade_date,
                    'total_record_count': record.total_record_count,
                }
            else:
                date_data[date] = {
                    'no_of_data_available': '-',
                    'no_of_data_scraped': '-',
                    'Scraped_on': '-',
                    'status': '-',
                    'Table Name': '-',
                    'Reason': '-',
                    'Trade Date': '-',
                    'total_record_count': '-',
                }

        # Append table data to the main structured_data dictionary
        structured_data[table_name] = date_data

    # Context data for rendering the HTML template
    context = {
        'table_names': table_names,
        'date_range': date_range,
        'structured_data': structured_data,
        'amber_table_names': amber_table_names,
        'failure_reasons': failure_reasons,
        'today': today,
        'start_date': start_date,
        'yesterday': yesterday,
        'table_list': table_list,
    }

    return render(request, 'pivottable.html', context)



def get_data_for_popup(request, table_name):
    today = datetime.today().date()
    yesterday = today - timedelta(days=1)

    data = scraping_log.objects.filter(table_name=table_name,trade_date=yesterday).first()

    amber_table_names = [
        'ace_52whl',
        'ace_company_master',
        'ace_Finance_fr',
        'ace_Finance_cf',
        'ace_company_equity',
        'ace_Finance_cons_cf',
        'ace_company_equity_cons',
        'ace_finance_Quarterly',
        'ace_finance_Quarterly_Cons',
        'ace_shp',
    ]

    if data:
        # Format the timestamp into a string with a specific format
        formatted_scraped_on = data.Scraped_on.strftime("%d-%m-%Y")
        # Assuming data.trade_date is a string in the format "2023-12-29"
        trade_date_str = data.trade_date

# Convert the string to a datetime object
        trade_date = datetime.strptime(trade_date_str, "%Y-%m-%d")

# Format the datetime object as "29-12-2023"
        formatted_trade_date = trade_date.strftime("%d-%m-%Y")

        response_data = {
            'table_name': data.table_name,
            'status': data.status,
            'no_of_data_scraped': data.no_of_data_scraped,
            'reason': data.reason,
            'amber_table_names': amber_table_names,
            'trade_date': formatted_trade_date,
            'Scraped_on': formatted_scraped_on,  # Use the formatted timestamp
        }
        return JsonResponse(response_data)
    else:
        return JsonResponse({'message': 'Data not available for today.'})
